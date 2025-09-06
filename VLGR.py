import os
import glob
import subprocess
from tqdm.notebook import trange
import pandas as pd
import re
from openpyxl import load_workbook
import calendar

def convert_and_replace_xls_to_xlsx(root_folder):
    """
    Рекурсивно конвертирует все .xls-файлы в указанной папке (и вложенных папках) в .xlsx с помощью LibreOffice.
    Новые .xlsx-файлы сохраняются в тех же папках. Исходные .xls-файлы удаляются.
    Требует установленной LibreOffice (на Colab - !apt-get install -y libreoffice).
    """
    # Проверка наличия libreoffice
    try:
        subprocess.run(['libreoffice', '--version'], check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    except Exception:
        raise RuntimeError('LibreOffice не установлена. Установите ее перед запуском этой функции!')

    # Рекурсивный поиск .xls
    xls_files = [y for x in os.walk(root_folder) for y in glob.glob(os.path.join(x[0], '*.xls'))]

    for i in trange(len(xls_files), desc="Конвертация xls → xlsx", unit="файл"):
        xls_path = xls_files[i]
        folder_path = os.path.dirname(xls_path)
        file_name = os.path.splitext(os.path.basename(xls_path))[0]
        xlsx_path = os.path.join(folder_path, file_name + '.xlsx')

        # Конвертация xls в xlsx
        try:
            subprocess.run([
                'libreoffice', '--headless', '--convert-to', 'xlsx', '--outdir', folder_path, xls_path
            ], check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            if os.path.exists(xlsx_path):
                os.remove(xls_path)
        except Exception as e:
            pass  # Можно добавить логгирование ошибок, если нужно

# Пример вызова:
# convert_and_replace_xls_to_xlsx('/content/gdrive/MyDrive/VLGR')



def excel_parser_STATEMENT(file_path):
    """
    Парсит Excel-файл в потоковый DataFrame.

    Параметры:
    - file_path: str, путь к файлу Excel

    Возвращает:
    - DataFrame с потоковой структурой данных
    """
    def extract_month_year(text):
        match = re.search(r'([А-ЯЁа-яё]+)\s+(\d{4})', text)
        return f"{match.group(1)} {match.group(2)}" if match else text

    def get_cell_color(cell):
        return cell.fill.start_color.rgb if cell.fill.start_color else None

    wb = load_workbook(file_path, data_only=True)
    sheet = wb.active

    # Автоматическое формирование маски из ячеек A6, A7
    level_names = {
        'account': sheet['A6'].value if sheet['A6'].value else None,
        'sublevel': sheet['A7'].value if sheet['A7'].value else None,
        'detail': sheet['A8'].value if sheet['A8'].value else None,
    }

    company_name = sheet['A1'].value.strip()
    date_info = extract_month_year(sheet['A2'].value.strip())

    start_row = 9
    max_row = sheet.max_row

    columns_mapping = {
        2: ('Сальдо на начало периода', 'Дебет'),
        3: ('Сальдо на начало периода', 'Кредит'),
        4: ('Обороты за период', 'Дебет'),
        5: ('Обороты за период', 'Кредит'),
        6: ('Сальдо на конец периода', 'Дебет'),
        7: ('Сальдо на конец периода', 'Кредит'),
    }

    current_account = None
    current_sublevel = None
    rows_data = []

    for row in range(start_row, max_row + 1):
        cell = sheet.cell(row=row, column=1)
        cell_value = cell.value
        cell_color = get_cell_color(cell)

        if cell_color == 'FFD6E5CB' and (cell_value is not None and 'итого' in str(cell_value).strip().lower()):
            for col_idx in range(2, 8):
                cell_data = sheet.cell(row=row, column=col_idx).value
                if cell_data not in (None, ''):
                    indicator, debit_credit = columns_mapping[col_idx]
                    rows_data.append({
                        'Company': company_name,
                        'Period': date_info,
                        level_names['account']: cell_value,  # сохраняем точное название итога
                        level_names['sublevel']: None,
                        level_names['detail']: None,
                        'Показатель': indicator,
                        'Дебет/Кредит': debit_credit,
                        'Value': cell_data
                    })
            continue

        # if cell_color == 'FFE4F0DD':
        #     current_account = cell_value
        #     current_sublevel = None
            
        if cell_color == 'FFE4F0DD':
            current_account = cell_value
            current_sublevel = None
            # Добавляем агрегатную строку, если есть значения!
            for col_idx in range(2, 8):
                cell_data = sheet.cell(row=row, column=col_idx).value
                if cell_data not in (None, ''):
                    indicator, debit_credit = columns_mapping[col_idx]
                    rows_data.append({
                        'Company': company_name,
                        'Period': date_info,
                        level_names['account']: current_account,
                        level_names['sublevel']: None,
                        level_names['detail']: None,
                        'Показатель': indicator,
                        'Дебет/Кредит': debit_credit,
                        'Value': cell_data
                    })
            continue

        elif cell_color == 'FFF0F6EF':
            current_sublevel = cell_value

        elif cell_color == 'FFD6E5CB':
            continue

        else:
            for col_idx in range(2, 8):
                cell_data = sheet.cell(row=row, column=col_idx).value
                if cell_data not in (None, ''):
                    indicator, debit_credit = columns_mapping[col_idx]
                    rows_data.append({
                        'Company': company_name,
                        'Period': date_info,
                        level_names['account']: current_account,
                        level_names['sublevel']: current_sublevel,
                        level_names['detail']: cell_value,
                        'Показатель': indicator,
                        'Дебет/Кредит': debit_credit,
                        'Value': cell_data
                    })

    df = pd.DataFrame(rows_data)

    # -----------------------------
    if 'Счет, Наименование счета' in df.columns:
        df = df.rename(columns={'Счет, Наименование счета': 'Счет'})
        df['Счет'] = df['Счет'].astype(str).str.split(',', n=1).str[0].str.strip()
    # -----------------------------

    # После создания df
    # 1. Определяем столбцы уровней
    account_col = level_names['account']
    sublevel_col = level_names['sublevel']
    detail_col = level_names['detail']
    
    # 2. Если detail_col нет (None или пустой), а sublevel_col есть — ищем столбец без имени
    if not detail_col and sublevel_col:
        # Находим столбец с пустым заголовком (None или ''), если он есть
        empty_cols = [col for col in df.columns if not col]
        if empty_cols:
            empty_col = empty_cols[0]
            # Переносим данные в sublevel_col
            df[sublevel_col] = df[sublevel_col].combine_first(df[empty_col])
            # Удаляем пустой столбец
            df = df.drop(columns=[empty_col])
    
    def next_month_date(period_text):
        months = {
            'январь': 1, 'февраль': 2, 'март': 3, 'апрель': 4, 'май': 5, 'июнь': 6,
            'июль': 7, 'август': 8, 'сентябрь': 9, 'октябрь': 10, 'ноябрь': 11, 'декабрь': 12
        }
        period_text_clean = str(period_text).strip().replace('\xa0', ' ')
        period_lower = period_text_clean.lower()
        match = re.search(r'([а-яё]+)\s*(\d{4})', period_lower)
        if match:
            month_name = match.group(1)
            year = int(match.group(2))
            month = months.get(month_name)
            if month:
                if month == 12:
                    next_month = 1
                    next_year = year + 1
                else:
                    next_month = month + 1
                    next_year = year
                return f"{next_year}-{next_month:02d}-01"
        return period_text
    
    df['Period'] = pd.to_datetime(df['Period'].apply(next_month_date))

    # Универсальное переименование столбцов по словарю
    rename_dict = {
        'Контрагенты': 'Partner',
        'Договоры': 'Contract',
        'Подразделение': 'Estate',
        'Статьи движения денежных средств': 'Category',
        'Статьи затрат': 'Category',
        'Банковские счета': 'Bank Account',
        'Period': 'Date'
    }
    df = df.rename(columns=rename_dict)

    # 1. Добавить столбец если его нет
    if 'Category' not in df.columns:
        df['Category'] = None
    if 'Type' not in df.columns:
        df['Type'] = None
    if 'Document' not in df.columns:
        df['Document'] = None
    
    # 2. Перенести все значения из "Счет" с "итого" в "Category", а в "Счет" — оставить пусто
    if 'Счет' in df.columns:
        mask_itogo = df['Счет'].astype(str).str.lower().str.contains('итого', na=False)
        unique_accounts = df.loc[~mask_itogo, 'Счет'].dropna().astype(str).unique()
        # Если одно уникальное — используем его
        if len(unique_accounts) == 1:
            account_value = unique_accounts[0]
        # Если среди уникальных есть '76' — используем его
        elif '76' in unique_accounts:
            account_value = '76'
        # Если ни одно из условий не сработало — пусто
        else:
            account_value = None
        # Заполняем для строк "итого"
        df.loc[mask_itogo, 'Category'] = df.loc[mask_itogo, 'Счет']
        df.loc[mask_itogo, 'Счет'] = account_value

    # Желаемый порядок столбцов --------------------------------------
    desired_order = [
        'Date', 'Company', 'Estate', 'Type', 'Category', 
        'Partner', 'Contract', 'Document', 'Bank Account', 'Value'
    ]
    # Сначала берем те, которые есть, в нужном порядке
    columns_in_order = [col for col in desired_order if col in df.columns]
    # Потом добавляем остальные, которых нет в последовательности
    other_columns = [col for col in df.columns if col not in columns_in_order]
    # Итоговый порядок
    final_order = columns_in_order + other_columns
    # Переупорядочиваем DataFrame
    df = df[final_order]

    # Удаляем лишние пробелы
    def strip_and_normalize_spaces(df, columns):
        """
        Очищает пробелы (в начале/конце и внутри) во всех указанных столбцах df.
        None/NaN значения остаются пропущенными!
        """
        def clean_value(val):
            if pd.isnull(val):
                return val
            s = str(val).strip()
            s = re.sub(r'\s+', ' ', s)
            # Если после чистки осталась пустая строка, вернуть None (по желанию)
            if s.lower() in ['nan', 'none', '']:
                return None
            return s
        for col in columns:
            if col in df.columns:
                df[col] = df[col].apply(clean_value)
        return df
    strip_and_normalize_spaces(df, ['Estate', 'Category', 'Contract', 'Bank Account'])

    # Замена значений Category для корректного соответствия
    replace_dict = {
        'Аренда помещения': 'Аренда помещений'
    }
    df['Category'] = df['Category'].replace(replace_dict)
    
    return df

# # Пример использования
# df = excel_parser_STATEMENT('/content/ОСВ 76 февраль 2025.xlsx')
# df






import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook

def excel_parser_INCOME(file_path):
    """
    Парсит Excel-файл с анализом выручки в потоковую таблицу.
    Теперь поддерживает множественные оттенки цвета для секций, компаний и объектов.
    """

    from datetime import datetime, timedelta
    from openpyxl import load_workbook
    import pandas as pd

    def get_cell_color(cell):
        return cell.fill.fgColor.rgb if hasattr(cell.fill.fgColor, 'rgb') else None

    def get_next_month_firstday(date_range_str):
        if not isinstance(date_range_str, str):
            return None
        parts = date_range_str.split('-')
        if len(parts) < 2:
            return None
        date_str = parts[1].strip().split(' ')[0]
        try:
            dt = datetime.strptime(date_str, '%d.%m.%Y')
            next_month = (dt.replace(day=1) + timedelta(days=32)).replace(day=1)
            return next_month.strftime('%Y-%m-%d')
        except Exception:
            return None

    # ВАРИАНТЫ ЦВЕТОВ (добавьте сюда все оттенки, которые реально встречаются)
    SECTION_COLORS = ['00E0FFE0', 'FFE0FFE0', 'FFCCFFCC', '00CCFFCC', '00CFFFD7', None]
    COMPANY_COLORS = ['00A6CAF0', 'FFA6CAF0', 'FFB7DEE8', 'FFB7DEE9', None]
    OBJECT_COLORS  = ['00C0DCC0', 'FFC0DCC0', 'FF99CC99', 'FF92D050', 'FF00B050', None]

    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    report_date = ws['B3'].value.strip() if ws['B3'].value else None

    start_row = None
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=2).value and 'Наименование' in str(ws.cell(row=row, column=2).value):
            start_row = row + 1
            break
    if start_row is None:
        raise ValueError("Не найдена строка с заголовком 'Наименование'.")

    current_section = None
    current_company = None
    current_object = None
    rows_data = []

    for row in range(start_row, ws.max_row + 1):
        cell = ws.cell(row=row, column=2)
        cell_value = cell.value
        cell_color = get_cell_color(cell)

        # Сравниваем по спискам цветов:
        if cell_color in SECTION_COLORS:
            current_section = str(cell_value).strip() if cell_value else None
            current_company = None
            current_object = None
            continue
        elif cell_color in COMPANY_COLORS:
            current_company = str(cell_value).strip() if cell_value else None
            current_object = None
            continue
        elif cell_color in OBJECT_COLORS:
            current_object = str(cell_value).strip() if cell_value else None
            continue

        # Пропуск пустых строк
        if not any([ws.cell(row=row, column=col).value for col in range(2, 6)]):
            continue

        # Основные поля (2-5 столбцы)
        act_value        = ws.cell(row=row, column=2).value
        contract_value   = ws.cell(row=row, column=3).value
        contragent_value = ws.cell(row=row, column=4).value
        revenue_value    = ws.cell(row=row, column=5).value

        rows_data.append({
            'Date': report_date,
            'Category': current_section,
            'Company': current_company,
            'Estate': current_object,
            'Document': act_value,
            'Contract': contract_value,
            'Partner': contragent_value,
            'Value': revenue_value,
        })

    df = pd.DataFrame(rows_data)

    df['Date'] = df['Date'].apply(get_next_month_firstday)
    df['Date'] = pd.to_datetime(df['Date'])

    mask_itogo = (df['Document'] == 'Итого:')
    df.loc[mask_itogo, 'Category'] = 'Итого за месяц'
    df.loc[mask_itogo, ['Company', 'Estate', 'Document']] = None

    df = df.dropna(subset=['Company', 'Document', 'Partner', 'Value'], how='all').reset_index(drop=True)

    df['Type'] = "Доходы"
    df = df[['Date', 'Company', 'Estate', 'Type', 'Category', 'Partner', 'Contract', 'Document', 'Value']]
    df['Value'] = (
        df['Value']
        .astype(str)
        .str.replace(' ', '', regex=False)
        .str.replace(',', '.', regex=False)
        .replace('nan', None)
    )
    df['Value'] = pd.to_numeric(df['Value'], errors='coerce')

    # Если Company содержит только одно уникальное значение (не пустое), подставить его вместо пустых
    unique_companies = df['Company'].dropna().unique()
    if len(unique_companies) == 1:
        single_company = unique_companies[0]
        df['Company'] = df['Company'].fillna(single_company)

    df['Счет'] = 'Данные по выручке'
    
    # Удаляем лишние пробелы
    def strip_and_normalize_spaces(df, columns):
        """
        Очищает пробелы (в начале/конце и внутри) во всех указанных столбцах df.
        None/NaN значения остаются пропущенными!
        """
        def clean_value(val):
            if pd.isnull(val):
                return val
            s = str(val).strip()
            s = re.sub(r'\s+', ' ', s)
            # Если после чистки осталась пустая строка, вернуть None (по желанию)
            if s.lower() in ['nan', 'none', '']:
                return None
            return s
        for col in columns:
            if col in df.columns:
                df[col] = df[col].apply(clean_value)
        return df
    strip_and_normalize_spaces(df, ['Estate', 'Category', 'Contract', 'Document'])
    
    # Замена значений Category для корректного соответствия
    replace_dict = {
        'Аренда помещения': 'Аренда помещений'
    }
    df['Category'] = df['Category'].replace(replace_dict)
    
    return df



import os
import glob
import pandas as pd
from tqdm import trange

def parse_statement_folder(root_main, root_statement, parser_func):
    """
    Обрабатывает все .xlsx-файлы ОСВ/ведомостей во вложенных папках, объединяет их в единый DataFrame.

    Аргументы:
        root_main     (str): Корневая папка, например '/content/gdrive/MyDrive/Волгоград'
        root_statement(str): Подкаталог ведомостей, например 'Ведомость'
        parser_func (callable): Функция-парсер одного файла (например, VLGR.excel_parser_STATEMENT)

    Возвращает:
        pd.DataFrame: Общий потоковый DataFrame по всем найденным ведомостям
    """
    base_dir = os.path.join(root_main, root_statement)
    all_files = glob.glob(os.path.join(base_dir, '**', '*.xlsx'), recursive=True)

    print(f'Найдено файлов: {len(all_files)}')

    all_data = []
    for i in trange(len(all_files), desc="Парсинг файлов", unit="файл"):
        file = all_files[i]
        try:
            df = parser_func(file)
            # Получаем относительный путь файла от root_main:
            rel_path = os.path.relpath(file, root_main)
            df['SOURCE_FILE'] = rel_path
            # df['SOURCE_FILE'] = os.path.basename(file)  # если нужно имя файла
            all_data.append(df)
        except Exception as e:
            print(f'Ошибка при парсинге файла {file}: {e}')

    if all_data:
        df_all = pd.concat(all_data, ignore_index=True)
    else:
        df_all = pd.DataFrame()

    # Желаемый порядок столбцов --------------------------------------
    desired_order = [
        'Date', 'Company', 'Estate', 'Type', 'Category', 
        'Partner', 'Contract', 'Document', 'Bank Account', 'Value'
    ]

    df = df_all
    # Сначала берем те, которые есть, в нужном порядке
    columns_in_order = [col for col in desired_order if col in df.columns]
    # Потом добавляем остальные, которых нет в последовательности
    other_columns = [col for col in df.columns if col not in columns_in_order]
    # Итоговый порядок
    final_order = columns_in_order + other_columns
    # Переупорядочиваем DataFrame
    df = df[final_order]
    
    return df


def parse_income_folder(root_main, root_income, parser_func):
    """
    Проходит по всем .xlsx-файлам во вложенных каталогах, парсит их заданной функцией и объединяет в один DataFrame.

    Аргументы:
        root_main   (str): Корневая папка, например '/content/gdrive/MyDrive/Волгоград'
        root_income (str): Подкаталог выручки, например 'Выручка'
        parser_func (callable): Функция-парсер одного файла (например, VLGR.excel_parser_INCOME)

    Возвращает:
        pd.DataFrame: Общий потоковый DataFrame по всем найденным выгрузкам
    """

    # Формируем абсолютный путь к каталогу с выгрузками
    base_dir = os.path.join(root_main, root_income)

    # Рекурсивный поиск всех .xlsx файлов во всех вложенных папках
    all_files = glob.glob(os.path.join(base_dir, '**', '*.xlsx'), recursive=True)

    print(f'Найдено файлов: {len(all_files)}')

    all_data = []

    # Проходим по всем найденным файлам с прогресс-баром
    for i in trange(len(all_files), desc="Парсинг файлов", unit="файл"):
        file = all_files[i]
        try:
            df = parser_func(file)
            # Получаем относительный путь файла от root_main:
            rel_path = os.path.relpath(file, root_main)
            df['SOURCE_FILE'] = rel_path
            # df['SOURCE_FILE'] = os.path.basename(file)  # можно добавить имя файла в итоговую таблицу
            all_data.append(df)
        except Exception as e:
            print(f'Ошибка при парсинге файла {file}: {e}')

    # Объединяем все DataFrame в один
    if all_data:
        df_all = pd.concat(all_data, ignore_index=True)
    else:
        df_all = pd.DataFrame()

    return df_all











def normalize_company_name(name):
    opf_list = ['АНО ДПО', 'ООО', 'ЗАО', 'ОАО', 'НПО', 'АО', 'ПАО', 'ФГБУ', 'УФССП', 'УФФССП', 'КПК', 'ОСФР', 'УФК', 'НО', 'МУП']
    opf_pattern = '|'.join(sorted(opf_list, key=len, reverse=True))

    # Убираем лишние символы, стандартизируем кавычки
    name_clean = name.strip().replace('«', '"').replace('»', '"').replace('.', '. ').strip()
    name_clean = re.sub(r'\s+', ' ', name_clean)

    # --- НОВЫЙ ЭТАП: замена полных названий на сокращенные ОПФ ---
    replacements = {
        r'\bОбщество с ограниченной ответственностью\b': 'ООО'
    }
    for pattern, repl in replacements.items():
        name_clean = re.sub(pattern, repl, name_clean, flags=re.I)

    upper = name_clean.upper()

    # 1. ИП — CAPS, "ИП" в начале
    if re.search(r'\bИП\b', upper):
        ip_match = re.search(r'\bИП\b', upper)
        if ip_match.start() == 0:
            fio = name_clean[ip_match.end():].strip()
        else:
            fio = name_clean[:ip_match.start()].strip()
        fio = re.sub(r'\.', '', fio)
        fio = re.sub(r'\s+', ' ', fio).upper()
        result = f'ИП {fio}'

    # 2. Юрлицо с ОПФ — CAPS, ОПФ в начале
    elif (opf_match := re.search(r'\b(' + opf_pattern + r')\b', upper)):
        opf = opf_match.group(1)
        # Убираем ОПФ из исходного названия
        cleaned = re.sub(r'\b(' + opf_pattern + r')\b', '', upper)
        cleaned = cleaned.replace('"', '').replace('.', ' ').strip()
        result = f'{opf} {cleaned}'

    else:
        # 3. Сокращенные ФИО (Фамилия И.О.)
        fio_match = re.match(r'^([А-ЯЁа-яё]+)\s+([А-ЯЁа-яё])\.\s*([А-ЯЁа-яё])\.$', name_clean)
        fio_match_single_dot = re.match(r'^([А-ЯЁа-яё]+)\s+([А-ЯЁа-яё])\.([А-ЯЁа-яё])\.$', name_clean)
        fio_match_separated = re.match(r'^([А-ЯЁа-яё]+)\s+([А-ЯЁа-яё])\.\s+([А-ЯЁа-яё])\.$', name_clean)

        if fio_match or fio_match_single_dot or fio_match_separated:
            fio_groups = fio_match or fio_match_single_dot or fio_match_separated
            surname = fio_groups.group(1).title()
            initials = f"{fio_groups.group(2).upper()}.{fio_groups.group(3).upper()}."
            result = f"{surname} {initials}"

        # 4. Сокращенные ФИО (Фамилия И.О без одной точки)
        elif (fio_match_alt := re.match(r'^([А-ЯЁа-яё]+)\s+([А-ЯЁа-яё])\.([А-ЯЁа-яё])$', name_clean)):
            surname = fio_match_alt.group(1).title()
            initials = f"{fio_match_alt.group(2).upper()}.{fio_match_alt.group(3).upper()}."
            result = f"{surname} {initials}"

        # 5. Фамилия с одной буквой-инициалом (Фамилия И.)
        elif (single_init_match := re.match(r'^([А-ЯЁа-яё]+)\s+([А-ЯЁа-яё])\.$', name_clean)):
            surname = single_init_match.group(1).title()
            initial = f"{single_init_match.group(2).upper()}."
            result = f"{surname} {initial}"

        # 6. Полное ФИО (2+ слов) каждое слово с большой буквы
        elif len(name_clean.split()) >= 2:
            result = name_clean.title()

        # 7. Все остальные случаи — CAPS
        else:
            result = upper

    # Финальная чистка пробелов
    result = re.sub(r'\s+', ' ', result).strip()

    return result

def normalize_company_names(val):
    if pd.isnull(val) or str(val).strip() == None:
        return None
    return normalize_company_name(str(val))





















import os, re, glob
import pandas as pd
from openpyxl import load_workbook
from tqdm import trange

def excel_parser_SUPPLIERS(file_path: str, debug: bool=False) -> pd.DataFrame:
    """
    Парсер 'Поставщики услуг' с корректным разделением Счет/Value и разбиением Doc/AnDT/AnCR на списки.

    Логика колонок:
      • Определяем по шапке блоки 'Дебет/Дт' и 'Кредит/Кт'. Под каждым ищем подзаголовок 'Счет'.
        DtAccountCol = колонка 'Счет' под Дебетом; DtSumCol = DtAccountCol + 1.
        CrAccountCol = колонка 'Счет' под Кредитом; CrSumCol = CrAccountCol + 1.
      • Дата — из первого столбца.
      • Doc / AnDT / AnCR — разбиваются в СПИСКИ по переводам строк в ячейке; элементы '<...>' удаляются.
      • Обычные строки: 'Счет' — текст из AccountCol; 'Value' — число из SumCol.
      • Итоговые строки ('итог/обороты/сальдо'): если сумма «слилась» в AccountCol и SumCol пуст,
        переносим число в Value, 'Счет' оставляем None, НО только если содержимое AccountCol не похоже на код счёта.
      • Коды счётов распознаются по ^\\d{1,3}(\\.\\d{1,2})?$
    """

    # ----------------- утилиты -----------------
    def _cell_str(v):
        if v is None:
            return None
        s = str(v).strip()
        return s if s else None

    def _to_number(x):
        if x is None:
            return None
        if isinstance(x, (int, float)):
            return None if pd.isna(x) else float(x)
        s = str(x).replace('\xa0','').replace(' ','').replace(',','.')
        try:
            return float(s)
        except Exception:
            return None

    def _to_datetime(v):
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return None
        return pd.to_datetime(v, dayfirst=True, errors='coerce')

    def _looks_like_account(text) -> bool:
        """Распознаём коды счётов: 26, 51, 60.01, 76.09, 101, 101.02 и т.п."""
        if text is None:
            return False
        s = str(text).strip().replace('\xa0','').replace(' ','')
        return re.match(r'^\d{1,3}(?:\.\d{1,2})?$', s) is not None

    def _format_account_text(v):
        """Гарантируем текст для 'Счет' (не число, без экспонент, без лишних нулей)."""
        s = _cell_str(v)
        if s is not None:
            return s
        if isinstance(v, (int, float)) and not pd.isna(v):
            s = ('%f' % float(v)).rstrip('0').rstrip('.')
            return s
        return None

    def _is_total_context(values):
        """Определяем итоговый/оборотный контекст строки по ключевым словам."""
        text = ' '.join([x for x in values if x]).lower()
        return any(k in text for k in ('итог', 'обороты', 'оборот', 'сальдо'))

    def _split_cell_to_list(v):
        """
        Делит содержимое ячейки на список по переводам строк.
        Чистит пробелы, выбрасывает пустые элементы и точное значение '<...>'.
        """
        if v is None:
            return []
        s = str(v).replace('\r\n', '\n').replace('\r', '\n')
        parts = [p.strip() for p in s.split('\n')]
        parts = [p for p in parts if p and p != '<...>']
        return parts

    def _clean_text(x):
        if x is None or (isinstance(x, float) and pd.isna(x)):
            return None
        s = re.sub(r'\s+', ' ', str(x)).strip()
        return s if s else None

    # --------- поиски в листе (компания, старт, шапка) ----------
    def _find_company(ws):
        for r in range(1, min(ws.max_row, 4)+1):
            for c in range(1, 3+1):
                val = _cell_str(ws.cell(row=r, column=c).value)
                if val and len(val) >= 3 and not re.search(r'(период|отчет|дата|счет|наименование|организац)', val, re.I):
                    return val
        return _cell_str(ws['A1'].value) or _cell_str(ws['B1'].value) or _cell_str(ws['C1'].value)

    def _find_start_row(ws):
        max_r = min(ws.max_row, 120)
        max_c = min(ws.max_column, 50)
        for r in range(1, max_r+1):
            for c in range(1, max_c+1):
                v = _cell_str(ws.cell(row=r, column=c).value)
                if v and 'сальдо на начало' in v.lower():
                    return r + 1
        return 10  # дефолтно после шапки

    def _detect_columns_by_header(ws, start_row):
        """
        Ищем ячейки 'Дебет/Дт' и 'Кредит/Кт' и строго под ними 'Счет'.
        Возвращает (dt_acc_col, dt_sum_col, cr_acc_col, cr_sum_col) — 1-based.
        """
        max_c = min(ws.max_column, 50)
        head_bot = min(start_row + 3, ws.max_row)

        dt_acc_col = cr_acc_col = None

        def _is_debet(x):  return bool(x) and re.search(r'\b(дебет|дт)\b', x, re.I)
        def _is_credit(x): return bool(x) and re.search(r'\b(кредит|кт)\b', x, re.I)

        for r in range(1, head_bot + 1):
            for c in range(1, max_c + 1):
                val = _cell_str(ws.cell(row=r, column=c).value)
                if _is_debet(val):
                    below = _cell_str(ws.cell(row=r+1, column=c).value)
                    if below and 'счет' in below.lower():
                        dt_acc_col = c
                if _is_credit(val):
                    below = _cell_str(ws.cell(row=r+1, column=c).value)
                    if below and 'счет' in below.lower():
                        cr_acc_col = c

        if dt_acc_col is None:
            dt_acc_col = 5
        if cr_acc_col is None:
            cr_acc_col = 7

        if debug:
            print(f"[HEADER] DtAccountCol={dt_acc_col}, DtSumCol={dt_acc_col+1}, CrAccountCol={cr_acc_col}, CrSumCol={cr_acc_col+1}")

        return dt_acc_col, dt_acc_col + 1, cr_acc_col, cr_acc_col + 1

    # ----------------- основная логика -----------------
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    company   = _find_company(ws)
    start_row = _find_start_row(ws)
    dt_acc_col, dt_sum_col, cr_acc_col, cr_sum_col = _detect_columns_by_header(ws, start_row)

    out = []
    max_r = ws.max_row

    for r in range(start_row, max_r+1):
        # базовые поля
        c1 = ws.cell(row=r, column=1).value   # Date
        c2 = ws.cell(row=r, column=2).value   # Doc
        c3 = ws.cell(row=r, column=3).value   # AnDT
        c4 = ws.cell(row=r, column=4).value   # AnCR

        # детектированные колонки
        dt_acc_cell = ws.cell(row=r, column=dt_acc_col).value
        dt_sum_cell = ws.cell(row=r, column=dt_sum_col).value
        cr_acc_cell = ws.cell(row=r, column=cr_acc_col).value
        cr_sum_cell = ws.cell(row=r, column=cr_sum_col).value

        # пропускаем пустые строки
        if all(_cell_str(v) is None for v in [c1,c2,c3,c4,dt_acc_cell,dt_sum_cell,cr_acc_cell,cr_sum_cell]):
            continue

        # дата
        dt = _to_datetime(c1)
        date_out = None if (dt is None or pd.isna(dt)) else dt.date()

        # МНОГОСТРОЧНЫЕ ПОЛЯ -> СПИСКИ
        doc_list  = _split_cell_to_list(c2)
        andt_list = _split_cell_to_list(c3)
        ancr_list = _split_cell_to_list(c4)

        # контекст для «Итого/Обороты/Сальдо»
        context_vals = [
            _cell_str(ws.cell(row=r, column=k).value)
            for k in range(1, min(ws.max_column, 6)+1)
        ]
        is_total = _is_total_context(context_vals)

        # ---- ДЕБЕТ ----
        dt_acc_text_raw = _format_account_text(dt_acc_cell)
        dt_val_right    = _to_number(dt_sum_cell)

        dt_val = dt_val_right
        dt_acc_text = dt_acc_text_raw
        if dt_val is None and is_total and not _looks_like_account(dt_acc_text_raw):
            maybe_num = _to_number(dt_acc_cell)
            if maybe_num is not None and _cell_str(dt_sum_cell) is None:
                dt_val = maybe_num
                dt_acc_text = None

        if dt_val is not None:
            out.append({
                'Date'   : date_out,
                'Company': company,
                'Doc'    : doc_list,
                'AnDT'   : andt_list,
                'AnCR'   : ancr_list,
                'DtCr'   : 'Dt',
                'Счет'   : dt_acc_text,
                'Value'  : dt_val
            })

        # ---- КРЕДИТ ----
        cr_acc_text_raw = _format_account_text(cr_acc_cell)
        cr_val_right    = _to_number(cr_sum_cell)

        cr_val = cr_val_right
        cr_acc_text = cr_acc_text_raw
        if cr_val is None and is_total and not _looks_like_account(cr_acc_text_raw):
            maybe_num = _to_number(cr_acc_cell)
            if maybe_num is not None and _cell_str(cr_sum_cell) is None:
                cr_val = maybe_num
                cr_acc_text = None

        if cr_val is not None:
            out.append({
                'Date'   : date_out,
                'Company': company,
                'Doc'    : doc_list,
                'AnDT'   : andt_list,
                'AnCR'   : ancr_list,
                'DtCr'   : 'Cr',
                'Счет'   : cr_acc_text,
                'Value'  : cr_val
            })

    if not out:
        return pd.DataFrame(columns=['Date','Company','Doc','AnDT','AnCR','DtCr','Счет','Value'])

    df = pd.DataFrame(out)

    # финальная очистка только строковых полей
    for col in ['Company','Счет']:
        if col in df.columns:
            df[col] = df[col].apply(_clean_text)

    # порядок столбцов
    return df[['Date','Company','Doc','AnDT','AnCR','DtCr','Счет','Value']]


def parse_suppliers_folder(root_main: str,
                           root_suppliers: str = 'Поставщики услуг',
                           parser_func = None) -> pd.DataFrame:
    """
    Рекурсивно парсит все .xlsx из подкаталога 'Поставщики услуг' и объединяет.
    Возвращает: Date, Company, Doc(list), AnDT(list), AnCR(list), DtCr, Счет, Value, SOURCE_FILE
    """
    if parser_func is None:
        parser_func = excel_parser_SUPPLIERS

    base_dir = os.path.join(root_main, root_suppliers)
    files = glob.glob(os.path.join(base_dir, '**', '*.xlsx'), recursive=True)
    print(f'Найдено файлов: {len(files)}')

    frames = []
    for i in trange(len(files), desc="Поставщики услуг: парсинг", unit="файл"):
        f = files[i]
        try:
            df = parser_func(f)
            df['SOURCE_FILE'] = os.path.relpath(f, root_main)
            frames.append(df)
        except Exception as e:
            print(f'Ошибка {f}: {e}')

    if not frames:
        return pd.DataFrame(columns=['Date','Company','Doc','AnDT','AnCR','DtCr','Счет','Value','SOURCE_FILE'])

    out = pd.concat(frames, ignore_index=True)
    # финальная раскладка
    desired = ['Date','Company','Doc','AnDT','AnCR','DtCr','Счет','Value','SOURCE_FILE']
    other = [c for c in out.columns if c not in desired]
    return out[desired + other]















import re
import pandas as pd

def enrich_suppliers_semantics(
    df_suppliers: pd.DataFrame,
    root_estate_dictionary: str,
    category_source_df: pd.DataFrame,
    category_source_col: str = "Category",
    debug: bool = False,
    show_progress: bool = True,
    progress_each: int = 500,
    normalize_company_fn=None
) -> pd.DataFrame:
    """
    Пост-обработка результатов VLGR.parse_suppliers_folder.

    Новые правила:
      • Финальная проверка 'компанийности' в AnDT/AnCR: элемент считается названием компании,
        если НЕ распознан по другим правилам и содержит НЕ БОЛЕЕ двух цифр.
      • Применяем нормализацию к Partner и Supplier (normalize_company_names из VLGR.py).
      • Для строк с перерасчётом долга (в Doc есть 'Корректировка долга' или 'Переуступка долга'):
           - AnCR -> Partner
           - AnDT -> Related Company
           - 'Переуступка долга...' из Doc -> Category
           - 'Корректировка долга...' из Doc -> Document
    """

    # -------- прогресс: инициализация --------
    try:
        from tqdm.auto import tqdm
    except Exception:
        tqdm = None

    # функция нормализации компаний
    if normalize_company_fn is None:
        # попробуем найти в модуле VLGR
        try:
            # если эта функция лежит в том же модуле, она попадёт в globals()
            _candidate = globals().get("normalize_company_names", None)
            normalize_company_fn = _candidate if callable(_candidate) else (lambda x: x)
        except Exception:
            normalize_company_fn = (lambda x: x)

    df = df_suppliers.copy()

    # ---------- утилиты ----------
    def norm(s: str | None) -> str:
        if s is None: return ""
        s = str(s).lower()
        s = s.replace("\xa0", " ")
        s = re.sub(r"[\t\r\n]+", " ", s)
        s = re.sub(r"[\"'`«»“”„]", "", s)
        s = re.sub(r"\s+", " ", s).strip()
        return s

    def split_list_cell(v) -> list[str]:
        if v is None: return []
        if isinstance(v, list): parts = v
        else:
            txt = str(v).replace("\r\n","\n").replace("\r","\n")
            parts = [p.strip() for p in txt.split("\n")]
        return [p for p in parts if p and p != "<...>"]

    def starts_with_long_digits(s: str) -> bool:
        return re.match(r"^\d{14,}", s) is not None  # bank account heuristic

    def digits_count(s: str) -> int:
        return sum(ch.isdigit() for ch in s)

    def contains_digits(s: str) -> bool:
        return any(ch.isdigit() for ch in s)

    def ratio(a: str, b: str) -> float:
        a2, b2 = norm(a), norm(b)
        if not a2 or not b2: return 0.0
        if a2 in b2 or b2 in a2: return 1.0
        aw, bw = set(a2.split()), set(b2.split())
        if not aw or not bw: return 0.0
        inter = len(aw & bw); base = min(len(aw), len(bw))
        return inter / base if base else 0.0

    def fuzzy_has_match(item: str, candidates: list[str], thr: float) -> bool:
        if not item: return False
        best = 0.0
        n_item = norm(item)
        for c in candidates:
            r = ratio(n_item, c)
            if r > best: best = r
            if best >= thr: break
        return best >= thr

    def startswith_any(s: str, prefixes: list[str]) -> bool:
        s2 = norm(s)
        return any(s2.startswith(norm(p)) for p in prefixes)

    # ---------- справочники ----------
    # Estate — из словаря объектов
    try:
        dict_df = pd.read_excel(root_estate_dictionary)
        cols = {c.lower(): c for c in dict_df.columns}
        col_src = next((cols[k] for k in cols if "исходное" in k and "наимен" in k), None)
        col_std = next((cols[k] for k in cols if "наимен" in k and "объект"  in k), None)
        estate_terms = []
        if col_src: estate_terms += [norm(x) for x in dict_df[col_src].dropna().astype(str)]
        if col_std: estate_terms += [norm(x) for x in dict_df[col_std].dropna().astype(str)]
        estate_terms = sorted(set([x for x in estate_terms if x]))
    except Exception as e:
        if debug: print(f"[enrich] Не удалось прочитать словарь объектов: {e}")
        estate_terms = []

    # Category — ИЗ ДРУГОЙ ТАБЛИЦЫ (общая база)
    if category_source_col in category_source_df.columns:
        category_terms = sorted(set([norm(x) for x in category_source_df[category_source_col].dropna().astype(str) if norm(x)]))
    else:
        category_terms = []
        if debug: print(f"[enrich] В category_source_df нет столбца '{category_source_col}'")

    contract_terms = [norm("договор"), norm("дог.")]
    document_terms = [norm(x) for x in ["Поступление","Акт","Накладная","УПД","Списание"]]

    # Спец-префиксы перерасчёта долга
    DOC_PREFIX_REASSIGN = ["Переуступка долга"]
    DOC_PREFIX_CORRECT  = ["Корректировка долга"]

    # ---------- подготовка выходных столбцов ----------
    for col in ["Partner","Supplier","Related Company","Category","Estate","Contract","Document","Bank Account","temp"]:
        if col not in df.columns: df[col] = None

    # для статистики прогресса
    stats = {"Partner":0,"Supplier":0,"Related":0,"Category":0,"Estate":0,"Contract":0,"Document":0,"Bank":0}

    n = len(df)
    use_tqdm = show_progress and ('tqdm' in globals() and tqdm is not None)
    pbar = tqdm(total=n, desc="Enrich suppliers", mininterval=0.5) if use_tqdm else None

    # ---------- основной цикл ----------
    for idx in range(n):
        row = df.iloc[idx]
        dtcr = str(row.get("DtCr","")).strip()

        doc_items  = split_list_cell(row.get("Doc"))
        andt_items = split_list_cell(row.get("AnDT"))
        ancr_items = split_list_cell(row.get("AnCR"))

        partner = row.get("Partner")
        supplier = row.get("Supplier")
        related_company = row.get("Related Company")
        category = row.get("Category")
        estate = row.get("Estate")
        contract = row.get("Contract")
        document = row.get("Document")
        bank_account = row.get("Bank Account")

        leftovers = []

        # Определяем: это кейс перерасчёта долга?
        is_recalc_doc = any(startswith_any(x, DOC_PREFIX_REASSIGN + DOC_PREFIX_CORRECT) for x in doc_items)

        # Сразу обработаем спец-правила из Doc:
        #   'Переуступка долга...' -> Category
        #   'Корректировка долга...' -> Document
        for x in doc_items:
            if startswith_any(x, DOC_PREFIX_REASSIGN) and (category is None or str(category).strip() == ""):
                category = x; stats["Category"] += 1
            if startswith_any(x, DOC_PREFIX_CORRECT) and document is None:
                document = x; stats["Document"] += 1

        # Тэгируем всё для дальнейшей классификации
        tagged = [("Doc",x) for x in doc_items] + [("AnDT",x) for x in andt_items] + [("AnCR",x) for x in ancr_items]

        # Базовые правила для каждого элемента
        for origin, item in tagged:
            base = item.strip()
            if not base or base == "<...>": continue
            nbase = norm(base)

            # 1) Банк. счёт
            if not bank_account and starts_with_long_digits(nbase):
                bank_account = base; stats["Bank"] += 1; continue

            # 2) Документ (общий список)
            #    (спец-док 'Корректировка долга' уже обработан выше)
            if not document and fuzzy_has_match(nbase, document_terms, thr=0.8):
                document = base; stats["Document"] += 1; continue

            # 3) Договор
            if not contract and fuzzy_has_match(nbase, contract_terms, thr=0.8):
                contract = base; stats["Contract"] += 1; continue

            # 4) Объект
            if not estate and estate_terms and fuzzy_has_match(nbase, estate_terms, thr=0.7):
                estate = base; stats["Estate"] += 1; continue

            # 5) Категория (из внешнего справочника)
            if (category is None or str(category).strip() == "") and category_terms and fuzzy_has_match(nbase, category_terms, thr=0.75):
                category = base; stats["Category"] += 1; continue

            # на финальную проверку
            leftovers.append((origin, base))

        # -------- ФИНАЛЬНОЕ распределение названий компаний --------
        # Новое правило: «компанийность» = НЕ распознано и количество цифр <= 2
        def company_candidates(items):
            return [txt for txt in items if digits_count(norm(txt)) <= 2]

        # Из leftovers выделим кандидатов для AnDT/AnCR
        andt_left = [txt for (orig, txt) in leftovers if orig == "AnDT"]
        ancr_left = [txt for (orig, txt) in leftovers if orig == "AnCR"]

        andt_names = company_candidates(andt_left)
        ancr_names = company_candidates(ancr_left)

        if is_recalc_doc:
            # Особый режим перерасчёта долга:
            #   AnCR -> Partner, AnDT -> Related Company (игнорируем DtCr)
            if not partner and ancr_names:
                partner = ancr_names[0]; stats["Partner"] += 1
            if not related_company and andt_names:
                related_company = andt_names[0]; stats["Related"] += 1
        else:
            # Обычный режим:
            if dtcr == "Dt":
                # Дт-строка: AnDT -> Supplier, AnCR -> Partner
                if not supplier and andt_names:
                    supplier = andt_names[0]; stats["Supplier"] += 1
                if not partner and ancr_names:
                    partner = ancr_names[0]; stats["Partner"] += 1
            elif dtcr == "Cr":
                # Кт-строка: AnDT -> Partner, AnCR -> Supplier
                if not partner and andt_names:
                    partner = andt_names[0]; stats["Partner"] += 1
                if not supplier and ancr_names:
                    supplier = ancr_names[0]; stats["Supplier"] += 1

        # Уберём из leftovers то, что съедено как Partner/Supplier/Related Company
        consumed = set()
        if supplier:
            consumed.add(("AnDT", supplier)) if dtcr == "Dt" and not is_recalc_doc else consumed.add(("AnCR", supplier))
        if partner:
            consumed.add(("AnCR", partner)) if (dtcr == "Dt" and not is_recalc_doc) else consumed.add(("AnDT", partner))
        if related_company:
            consumed.add(("AnDT", related_company))
        temp_list = [txt for (orig, txt) in leftovers if (orig, txt) not in consumed]

        # --- Нормализация Partner/Supplier ---
        if partner is not None:
            partner = normalize_company_fn(partner)
        if supplier is not None:
            supplier = normalize_company_fn(supplier)

        # запись
        if partner is not None:          df.at[idx, "Partner"] = partner
        if supplier is not None:         df.at[idx, "Supplier"] = supplier
        if related_company is not None:  df.at[idx, "Related Company"] = related_company
        if estate is not None:           df.at[idx, "Estate"] = estate
        if contract is not None:         df.at[idx, "Contract"] = contract
        if document is not None:         df.at[idx, "Document"] = document
        if bank_account is not None:     df.at[idx, "Bank Account"] = bank_account
        if category is not None and (pd.isna(row.get("Category")) or str(row.get("Category")).strip() == ""):
            df.at[idx, "Category"] = category

        df.at[idx, "temp"] = temp_list

        # прогресс
        if use_tqdm:
            if (idx % progress_each) == 0:
                pbar.set_postfix(P=stats["Partner"], S=stats["Supplier"], R=stats["Related"],
                                 Cat=stats["Category"], Es=stats["Estate"], Ctr=stats["Contract"],
                                 Doc=stats["Document"], Bank=stats["Bank"])
            pbar.update(1)
        elif show_progress and (idx % progress_each) == 0:
            print(f"[{idx}/{n}] P={stats['Partner']} S={stats['Supplier']} R={stats['Related']} "
                  f"Cat={stats['Category']} Es={stats['Estate']} Ctr={stats['Contract']} "
                  f"Doc={stats['Document']} Bank={stats['Bank']}")

    if pbar is not None:
        pbar.set_postfix(P=stats["Partner"], S=stats["Supplier"], R=stats["Related"],
                         Cat=stats["Category"], Es=stats["Estate"], Ctr=stats["Contract"],
                         Doc=stats["Document"], Bank=stats["Bank"])
        pbar.close()

    if show_progress and not use_tqdm:
        print(f"[done {n}] P={stats['Partner']} S={stats['Supplier']} R={stats['Related']} "
              f"Cat={stats['Category']} Es={stats['Estate']} Ctr={stats['Contract']} "
              f"Doc={stats['Document']} Bank={stats['Bank']}")

    return df

